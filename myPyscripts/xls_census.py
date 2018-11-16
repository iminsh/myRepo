#! python3
# excel file operation
# Tabulates population and number of census tracts
# for each country
# read the data and do some statistic, and write the
# result in a .py file for other module to use

import openpyxl, pprint

print('Opening workbook...')
wb = openpyxl.load_workbook(
    'D:\\automate_online-materials\\censuspopdata.xlsx')

### sheet
sheet = wb['Population by Census Tract']
countryData = {}

# fill in contryData with each country's population and tracts.
print('Reading rows...')
MAX_ROW = sheet.max_row + 1
for row in range(2, MAX_ROW):
    state = sheet.cell(row, 2).value
    country = sheet.cell(row, 3).value
    pop = sheet.cell(row, 4).value
#    print(state, country, pop)

    countryData.setdefault(state, {})
    countryData[state].setdefault(country, {'tracts': 0, 'pop': 0})
    countryData[state][country]['tracts'] += 1
    countryData[state][country]['pop'] += pop

# open a new .py file and write the content of countryData to it
# so that other module can import it to use the data
print('Writing result...')
resultFile = open('census2010.py', 'w')
resultFile.write('census2010 = ' + pprint.pformat(countryData))
resultFile.close()

wb.close()
print('Done.')

# at the end, but not the end
# you can import the data to be used in other modules:
# from census2010 import census2010
# census2010['NY']['Niagara']
