#! python3
import sys
import openpyxl
from openpyxl.styles import Font

if (len(sys.argv)== 2):
    if (not sys.argv[1].isdecimal()):
        print('Usage: py multiTable.py <N>')
        exit(0)
else:
    print('Usage: py multiTable.py <N>')
    exit(0)

N = int(sys.argv[1])

font = Font(bold=True)

wb = openpyxl.Workbook()
sheet = wb.active

col = 1
for row in range(1, N+1):
    sheet.cell(row=row+1, column=col).value = row
    sheet.cell(row=row+1, column=col).font = font

row = 1
for col in range(1, N+1):
    sheet.cell(row=row, column=col+1).value = col
    sheet.cell(row=row, column=col+1).font = font
    
for row in range(1,(N+1)):
    for col in range(1,(N+1)):
        sheet.cell(row=row+1, column=col+1).value = row * col
wb.save('multiTable.xlsx')

print("Done.")    




                  
